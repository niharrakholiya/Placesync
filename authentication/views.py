
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.core.checks import messages
from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.urls import reverse
from django.contrib.auth.hashers import check_password
from .forms import  CompanyLoginForm
from .forms import StudentLoginForm
from django.contrib.auth.hashers import make_password
from .models import Company, JobPost, BasicUser, Student, JobApplication, ApplicationStatus
from django.shortcuts import render
from django.contrib.auth import logout
from django.shortcuts import redirect
from django.http import HttpResponse
from openpyxl import Workbook
# Create your views here.


def home_page(request):
    return render(request, "index.html")


def student_login(request):
    if request.method == 'POST':
        form = StudentLoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            print("User:", user)
            if user is not None:
                if check_password(password, user.password):  # Check hashed password
                    login(request, user)
                    if hasattr(user, 'student_profile'):
                     request.session['stdid'] = user.student_profile.id
                     return redirect('student-dashboard')
                    else:
                        messages.error(request, 'Invalid username or password.')
                else:
                    messages.error(request, 'Invalid username or password.')
            else:
                messages.error(request, 'Invalid username or password.')
        else:
            messages.error(request, 'Invalid form submission.')
    else:
        form = StudentLoginForm()
    return render(request, 'student-login.html', {'form': form})



def company_login(request):
    if request.method == 'POST':
        form = CompanyLoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                if hasattr(user, 'company_profile'):
                    request.session['userid'] = user.company_profile.id
                    return redirect('company-dashboard')
                else:
                    messages.error(request, 'Invalid username or password.')
            else:
                messages.error(request, 'Invalid username or password.')
        else:
            messages.error(request, 'Invalid form submission.')
    else:
        form = CompanyLoginForm()
    return render(request, 'company-login.html', {'form': form})


def register_student(request):
    if request.method == 'POST':
        # Retrieve form data
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        gender = request.POST.get('gender')
        age = request.POST.get('age')
        dob = request.POST.get('dob')  # Assuming dob is submitted in the format YYYY-MM-DD
        phone_number = request.POST.get('phone-number')
        address = request.POST.get('address')
        state = request.POST.get('state')
        pincode = request.POST.get('pincode')
        photo = request.FILES.get('myfile')
        resume = request.FILES.get('resume')
        cpi = request.POST.get('cpi')

        # Create BasicUser object using CustomUserManager and save to database
        basic_user = BasicUser.objects.create_user(username=username, email=email, password=password)

        # Create Student object and associate it with the newly created BasicUser
        student = Student.objects.create(
            user=basic_user,
            student_name=username,
            gender=gender,
            age=age,
            dob=dob,
            email=email,
            phone_number=phone_number,
            address=address,
            state=state,
            pincode=pincode,
            photo=photo,
            resume=resume,
            cpi=cpi
        )

        return redirect('student-login')  # Redirect to the login page after successful registration
    else:
        return render(request, 'student-register.html')  # Render the registration form template


def register_company(request):
    if request.method == 'POST':
        # Retrieve form data
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        company_name = request.POST.get('company_name')
        company_location = request.POST.get('company_location')
        company_link = request.POST.get('company_link')
        about = request.POST.get('about')
        company_image = request.FILES.get('company_image')

        # Get the custom user model
        User = get_user_model()

        # Create BasicUser object using CustomUserManager and save to database
        basic_user = User.objects.create_user(username=username, email=email, password=password, role='company')

        # Create Company object and associate it with the newly created BasicUser
        company = Company(
            user=basic_user,
            company_name=company_name,
            company_location=company_location,
            company_link=company_link,
            about=about,
            company_image=company_image
        )
        company.save()

        return redirect('company-login')
    else:
        return render(request, 'company-register.html')


@login_required(login_url='company-login')
def company_dashboard(request):

    user = request.user
    context = {}
    print(user)
    if hasattr(user, 'company_profile'):
        company = user.company_profile
        print(user.company_profile)
        print(company)  # Debug: Print company object
        recent_openings = JobPost.objects.filter(company=company)[:5]
        context = {
            'company': company,  # Pass the company object directly
            'recent_openings': recent_openings

        }
    return render(request, 'company-dashboard.html', context)


def companies(request):
    # Query all registered companies from the database
    # desired_image_url = None
    registered_companies = Company.objects.exclude(company_name="company")

    # Pass the registered companies to the template context
    context = {
        'registered_companies': registered_companies
    }

    # Render the HTML template with the provided context
    return render(request, 'companies.html', context)


def logout_user(request):
    logout(request)
    return redirect('home-page')


@login_required(login_url='company-login')
def add_job_post(request):
    if request.method == 'POST':
        company = Company.objects.get(id=request.session.get('userid'))
        # for data in company:
        #     print(data)
        print("company   ")
        print(company.id)
        salary = request.POST.get('salary')
        positions = request.POST.get('positions')
        location = request.POST.get('location')
        total_posts = request.POST.get('total_posts')

        print(salary, positions, location, total_posts)
        print("you are printing job posts dear")
        jobpost = JobPost(
            company=company,
            salary=salary,
            positions=positions,
            location=location,
            total_posts=total_posts
        )
        print(jobpost)

        try:
            jobpost.save()
            print("JobPost object saved successfully")
        except Exception as e:
            print(f"Error saving JobPost object: {e}")
        return redirect('company-dashboard')
    else:
        return render(request, 'job-post.html')


def jobs(request):
    # Retrieve recent job openings
    recent_openings = JobPost.objects.all()[:5]  # Retrieve the 5 most recent job openings
    context = {
        'recent_openings': recent_openings
    }
    return render(request, 'company-dashboard.html', context)


@login_required(login_url='student-login')
def job_list(request):
    # Retrieve all job openings
    job_openings = JobPost.objects.all()
    print(job_openings)
    context = {
        'job_openings': job_openings
    }
    return render(request, 'joblisting.html', context)


# views.py


@login_required(login_url='student-login')
def student_dashboard(request):
    student_id = request.session.get('stdid')

    if student_id is None:
        # Handle the case where student_id is not found in the session
        return redirect('student-login')

    try:
        student = Student.objects.get(id=student_id)
    except ObjectDoesNotExist:
        # Handle the case where no student with the given id exists
        return redirect('error-page')

    context = {
        'student': student
    }

    return render(request, 'student-dashboard.html', context)

from django.contrib import messages

def apply_job(request, job_id, company, position):
    if request.method == 'POST':
        # Retrieve student data from session or however it's stored
        student_id = request.session.get('stdid')
        student = Student.objects.get(id=student_id)

        # Check if the student has already applied for a job posted by the same company
        existing_applications = JobApplication.objects.filter(student=student, company=company)
        if existing_applications.exists():
            # Set error message
            messages.error(request, 'You have already applied for a job posted by this company.')
            return redirect('joblist')

        # Create job application instance
        job_application = JobApplication(
            job_id=job_id,
            student=student,
            company=company,
            position=position,
        )
        job_application.save()

        messages.success(request, 'Your application has been submitted successfully!')
        return redirect('joblist')
    else:
        messages.error(request, 'Invalid request method.')
        return redirect('student-login')


@login_required(login_url='company-login')
def job_retrive(request):
    user_id = request.session.get('userid')
    user = Company.objects.get(id=user_id)
    company_name = user.company_name

    # Get IDs of applications rejected by other companies
    rejected_by_others = ApplicationStatus.objects.filter(
        status=ApplicationStatus.REJECTED
    ).values_list(
        'application_id', flat=True
    )
    print(rejected_by_others)
    # Get IDs of applications accepted by the current company
    accepted_by_company = ApplicationStatus.objects.filter(
        status=ApplicationStatus.ACCEPTED,
        is_accepted=True  # Filter out already accepted applications
    ).values_list('application_id', flat=True)
    print(accepted_by_company)
    # Get all applications for the current company where students are not accepted by other companies
    accepted_students = ApplicationStatus.objects.filter(
        application_id__in=accepted_by_company
    ).values_list('student_name', flat=True)
    print(accepted_students)
    # Get student IDs corresponding to the accepted students
    accepted_student_ids = Student.objects.filter(
        student_name__in=accepted_students
    ).values_list('id', flat=True)
    print(accepted_student_ids)
    # Get job applications for the current company excluding accepted students
    remaining_applications = JobApplication.objects.filter(company=company_name).exclude(student_id__in=accepted_student_ids)
    return render(request, 'job_retrive.html', {'job_applications': remaining_applications})



@login_required(login_url='company-login')
def accept_reject_application(request, application_id):
    application = get_object_or_404(JobApplication, id=application_id)
    company_name = application.company
    student_name = application.student.student_name

    if request.method == 'POST':
        if 'accept' in request.POST:
            # Check if a status entry already exists for this application and update it
            status, created = ApplicationStatus.objects.update_or_create(
                application=application,
                defaults={'company_name': company_name,
                          'student_name': student_name,
                          'status': ApplicationStatus.ACCEPTED}
            )
            # Update is_accepted to 1
            status.is_accepted = True
            status.save()
        elif 'reject' in request.POST:
            # Check if a status entry already exists for this application and update it
            status, created = ApplicationStatus.objects.update_or_create(
                application=application,
                defaults={'company_name': company_name,
                          'student_name': student_name,
                          'status': ApplicationStatus.REJECTED}
            )
            # Update is_accepted to 0
            status.is_accepted = False
            status.save()

        return redirect('job_retrive')

    return render(request, 'job_retrive.html', {'application': application})

@login_required(login_url='company-login')
def accepted_students(request):
    # Assuming you have a variable `company_id` that holds the ID of the specific company
    company_id = request.session.get('userid')
    user = Company.objects.get(id=company_id)
    company_name = user.company_name
    # Retrieve accepted job applications for the specific company
    print(company_name)
    accepted_applications = ApplicationStatus.objects.filter(status='AC', application__company=company_name)
    print(accepted_applications)
    # Initialize a list to store accepted students' details
    accepted_students = []

    for application in accepted_applications:
        # Fetch email from the related student
        email = application.application.student.email
        student_name = application.application.student.student_name
        # Fetch position applied for from the application
        position = application.application.position

        # Append the details to the list of accepted students
        accepted_students.append({'email': email, 'position': position,'student_name': student_name})
        print(accepted_students)
    return render(request, 'view.html', {'accepted_students': accepted_students})

import pandas as pd
import io
from openpyxl import Workbook
from django.http import HttpResponse


from django.conf import settings

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
import io
import pandas as pd

def download_excel(request):
    # Fetch the data from the database for a specific company
    company_id = request.session.get('userid')
    user = Company.objects.get(id=company_id)
    company_name = user.company_name
    accepted_applications = ApplicationStatus.objects.filter(status='AC', company_name=company_name)

    # Extract the required fields from ApplicationStatus model and related Student model
    applications_data = [
        {
            'Company Name': company_name,
            'Student Name': app.application.student.student_name,
            'Gender': app.application.student.gender,
            'Age': app.application.student.age,
            'Date of Birth': app.application.student.dob,
            'Phone Number': app.application.student.phone_number,
            'Address': app.application.student.address,
            'State': app.application.student.state,
            'Pincode': app.application.student.pincode,
            'CPI': app.application.student.cpi,
            'Resume URL': request.build_absolute_uri(app.application.student.resume.url) if app.application.student.resume else '',
            'Photo URL': request.build_absolute_uri(app.application.student.photo.url) if app.application.student.photo else '',
            # Include other fields as needed
        }
        for app in accepted_applications
    ]

    # Create a DataFrame for the accepted applications data
    df = pd.DataFrame(applications_data)

    # Create a new Excel file in-memory using io.BytesIO
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Accepted Students'

    # Set the column names in the first row of the Excel file
    column_names = list(df.columns)
    for col_idx, column_name in enumerate(column_names, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=column_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Write the DataFrame to the Excel file
    for idx, row in df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            cell = worksheet.cell(row=idx + 2, column=col_idx, value=value)
            # Add hyperlinks for Resume URL and Photo URL
            if column_names[col_idx - 1] in ['Resume URL', 'Photo URL'] and value:
                cell.hyperlink = value

    # Set the appropriate content type and attachment headers for the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="accepted_students.xlsx"'

    # Save the Excel file to the in-memory BytesIO object
    workbook.save(output)

    # Seek to the beginning of the BytesIO object
    output.seek(0)

    # Set the content of the response to the content of the BytesIO object
    response.write(output.read())

    return response

def about(request):
    return render(request, 'about.html')


def stop_job_opening(request, job_post_id):
    job_post = JobPost.objects.get(id=job_post_id)
    job_post.active = False
    job_post.save()
    return redirect('company-dashboard')

